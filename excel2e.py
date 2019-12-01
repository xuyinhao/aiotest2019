from openpyxl import Workbook,load_workbook
import  os
class ReadExcel():
    def __init__(self,filepath=None,shsheet=None):
        self.sheetname = shsheet
        if filepath:
            configpath = filepath
        else:
            root_dir = os.path.dirname(os.path.abspath(__file__))
            configpath = os.path.join(root_dir,"testdata.xlsx")
        self.cf = load_workbook(filename=configpath)
    def get_row_totlenum(self, parm):
        login_sheet = self.cf[str(parm)]
        return login_sheet.max_row
    def get_column_totlenum(self,parm):
        login_sheet = self.cf[str(parm)]
        return login_sheet.max_column
    def get_rowcol_num(self,parm):
        login_sheet = self.cf[str(parm)]
        return (login_sheet.max_row , login_sheet.max_column)
    # def letterschange(self,colnum):
    #     strabc = 'ABCDEFGHIJ'
    #     return strabc[:colnum]
    def save_excel(self,fileSavePath):
        self.cf.save(fileSavePath)

    def getValue(self,sheetparm):
        # print(self.letterschange(columnnum))
        (max_rownum,max_colnum) = self.get_rowcol_num(sheetparm)
        a = [];b = [];
        self.login_sheet = self.cf[sheetparm]
        for now_row in range(2, max_rownum + 1):
            for now_col in range(1, max_colnum + 1):
                a.append(self.login_sheet.cell(row=now_row,column=now_col).value)
            b.append(tuple(a[:]))
            a=[]
        return b

    def get_row_column(self,r_row,r_column):
        self.now_sheet = self.cf[self.sheetname]
        return self.now_sheet.cell(row=r_row,column=r_column).value
    def set_row_column(self,w_row,w_column,writeValue):
        self.write_sheet = self.cf[self.sheetname]
        # print(w_row,w_column,writeValue)
        try:
            self.write_sheet.cell(row=w_row,column=w_column).value=writeValue
        except Exception as e:
            print(e)

if __name__ == '__main__':
    import time
    #获取excel的sheet句柄
    ex_new=ReadExcel(filepath=r'g:\new.xlsx',shsheet='new')
    ex_ys=ReadExcel(filepath=r'g:\ys.xlsx',shsheet='ys')
    # print(ex_new.get_row_totlenum('new'))
    #获取 总共的row个数
    ex_new_max_row=ex_new.get_row_totlenum('new')
    # ex_new_max_row=10
    ex_ys_max_row=ex_ys.get_row_totlenum('ys')
    #从第二行开始，如果new里面内容在 ys_now里面 就追加信息到ys表
    start_time=time.time()
    for new_row in range(2,ex_new_max_row+1):
        ex_new_name = ex_new.get_row_column(new_row,1)
        for ys_row in range(2,ex_ys_max_row+1):
            if ex_new_name in ex_ys.get_row_column(ys_row,3):
                huxing=ex_new.get_row_column(new_row,2)
                danjia=ex_new.get_row_column(new_row,4)
                name=ex_new_name
                ex_ys.set_row_column(ys_row,5,huxing)
                ex_ys.set_row_column(ys_row,6,danjia)
                ex_ys.set_row_column(ys_row,7,name)
                break
    ex_ys.save_excel(r'g:\ys_save2.xlsx')
    print(time.time()-start_time)


    #ex_new.get_row_column(1,1)
    #ex_ys.get_row_column(1,2)
 # 2,3,4 -> 5,6,7