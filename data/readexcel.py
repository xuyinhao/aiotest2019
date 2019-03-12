from openpyxl import Workbook,load_workbook
import  os
# wb = load_workbook(filename='testdata.xlsx')
# login_sheet = wb['login']
# print(login_sheet.max_row,login_sheet.max_column)
# for now_row in range(2,login_sheet.max_row+1):
#     print(login_sheet.cell(row=now_row,column=1).value)
#     print(login_sheet.cell(row=now_row,column=2).value)
#     print(login_sheet.cell(row=now_row,column=3).value)
#     print('ok')

class ReadExcel():
    def __init__(self,filepath=None):
        if filepath:
            configpath = filepath
        else:
            root_dir = os.path.dirname(os.path.abspath(__file__))
            configpath = os.path.join(root_dir,"testdata.xlsx")
        self.cf = load_workbook(filename=configpath)
    def get_row_totlenum(self, parm):
        login_sheet = self.cf[str(parm)]
        return login_sheet.max_row
    def get_column_totlenum(self,parm):
        login_sheet = self.cf[str(parm)]
        return login_sheet.max_column
    def get_rowcol_num(self,parm):
        login_sheet = self.cf[str(parm)]
        return (login_sheet.max_row , login_sheet.max_column)
    # def letterschange(self,colnum):
    #     strabc = 'ABCDEFGHIJ'
    #     return strabc[:colnum]

    def getValue(self,sheetparm):
        # print(self.letterschange(columnnum))
        (max_rownum,max_colnum) = self.get_rowcol_num(sheetparm)
        a = [];b = [];
        self.login_sheet = self.cf[sheetparm]
        for now_row in range(2, max_rownum + 1):
            for now_col in range(1, max_colnum + 1):
                a.append(self.login_sheet.cell(row=now_row,column=now_col).value)
            b.append(tuple(a[:]))
            a=[]
        return b


if __name__ == '__main__':
    test=ReadExcel()
    #print(serverhost)
    # print(test.getRowColNum('login'))
